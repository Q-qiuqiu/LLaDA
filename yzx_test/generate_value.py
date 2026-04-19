import argparse
import json
import math
import os
from typing import Dict, List, Optional, Tuple

import numpy as np
import torch
import torch.nn.functional as F
from openpyxl import Workbook, load_workbook
from transformers import AutoModel, AutoTokenizer


def add_gumbel_noise(logits: torch.Tensor, temperature: float) -> torch.Tensor:
    """Add Gumbel noise to logits."""
    if temperature == 0:
        return logits
    logits = logits.to(torch.float64)
    noise = torch.rand_like(logits, dtype=torch.float64)
    gumbel_noise = (-torch.log(noise)) ** temperature
    return logits.exp() / gumbel_noise


def get_num_transfer_tokens(mask_index: torch.Tensor, steps: int) -> torch.Tensor:
    """
    Precompute how many tokens should be transferred at each denoising step
    for each sample in a block (same as original generate.py logic).
    """
    mask_num = mask_index.sum(dim=1, keepdim=True)
    base = mask_num // steps
    remainder = mask_num % steps
    num_transfer_tokens = (
        torch.zeros(mask_num.size(0), steps, device=mask_index.device, dtype=torch.int64) + base
    )
    for i in range(mask_num.size(0)):
        num_transfer_tokens[i, : remainder[i]] += 1
    return num_transfer_tokens


def count_remaining_mask_ratio(x: torch.Tensor, prompt_len: int, mask_id: int) -> float:
    """Remaining mask ratio in generation region."""
    gen_region = x[:, prompt_len:]
    remain = (gen_region == mask_id).sum().item()
    total = gen_region.numel()
    return remain / max(total, 1)


def _parse_step_block_from_xlsx(sheet, start_row: int, seq_len: int) -> Tuple[np.ndarray, np.ndarray]:
    """
    Parse one step block in denoise_value xlsx.
    Returns:
      attn: [L, L]
      conf: [L]
    """
    attn_rows: List[List[float]] = []
    conf_vals: List[float] = []
    row = start_row
    for _ in range(seq_len):
        attn_json = sheet.cell(row=row, column=3).value
        conf_val = sheet.cell(row=row, column=4).value
        attn_rows.append(json.loads(attn_json) if isinstance(attn_json, str) else list(attn_json))
        conf_vals.append(float(conf_val))
        row += 1
    return np.array(attn_rows, dtype=np.float64), np.array(conf_vals, dtype=np.float64)


def _parse_delta_attn_block_from_xlsx(sheet, start_row: int, seq_len: int) -> np.ndarray:
    """
    Parse one delta block in denoise_value xlsx.
    Returns:
      delta_attn: [L, L]
    """
    delta_rows: List[List[float]] = []
    row = start_row
    for _ in range(seq_len):
        delta_json = sheet.cell(row=row, column=3).value
        delta_rows.append(json.loads(delta_json) if isinstance(delta_json, str) else list(delta_json))
        row += 1
    return np.array(delta_rows, dtype=np.float64)


def _write_value_sheet(
    wb: Workbook,
    sheet_name: str,
    attention_scores: np.ndarray,
    top_confidences: np.ndarray,
    top_token_ids: np.ndarray,
    attention_deltas: np.ndarray,
    confidence_deltas: np.ndarray,
    prompt_len: int,
    top_token_text_per_step: List[List[str]],
) -> None:
    ws = wb.create_sheet(sheet_name)

    def token_cell(step_idx: int, pos_idx: int) -> str:
        token_id = int(top_token_ids[step_idx, pos_idx])
        token_text = ""
        if step_idx < len(top_token_text_per_step) and pos_idx < len(top_token_text_per_step[step_idx]):
            token_text = str(top_token_text_per_step[step_idx][pos_idx])
        token_text = token_text.replace("\n", "\\n")
        return f"{token_text} (id={token_id})"

    num_steps = int(attention_scores.shape[0])
    seq_len = int(attention_scores.shape[1])

    for step_idx in range(num_steps):
        ws.append([f"Step {step_idx + 1}", "", "", ""])
        ws.append(["sequence_token_index", "token", "attention_scores", "confidence"])
        for pos_idx in range(seq_len):
            seq_token_index = prompt_len + pos_idx + 1
            ws.append(
                [
                    seq_token_index,
                    token_cell(step_idx, pos_idx),
                    json.dumps(attention_scores[step_idx, pos_idx, :].tolist(), ensure_ascii=False),
                    float(top_confidences[step_idx, pos_idx]),
                ]
            )
        ws.append(["", "", "", ""])

    num_deltas = int(attention_deltas.shape[0])
    for pair_idx in range(num_deltas):
        left = pair_idx + 1
        right = pair_idx + 2
        ws.append([f"Step {left}-{right} Delta", "", "", ""])
        ws.append(["sequence_token_index", "token", "attention_delta", "confidence_delta"])
        token_step_idx = pair_idx + 1
        for pos_idx in range(seq_len):
            seq_token_index = prompt_len + pos_idx + 1
            ws.append(
                [
                    seq_token_index,
                    token_cell(token_step_idx, pos_idx),
                    json.dumps(attention_deltas[pair_idx, pos_idx, :].tolist(), ensure_ascii=False),
                    float(confidence_deltas[pair_idx, pos_idx]),
                ]
            )
        ws.append(["", "", "", ""])


def compute_side_xlsx_from_value_xlsx(
    value_xlsx_path: str,
    side_xlsx_path: str,
    stability_lambda: float = 1000.0,
    local_beta: float = 0.005,
    value_sheet: str = "denoise_values_mask_all",
) -> str:
    """
    Compute temporal dependency matrix from value xlsx and save as side xlsx.
    Formula:
      A_sym_t(i,j) = (A_t(i,j) + A_t(j,i))/2
      M0(i,j) = sum_t w_t * A_sym_t(i,j) * exp(-lambda * |DeltaA_t(i,j)|)
      M(i,j) = M0(i,j) * exp(-beta * |i-j|)
      with linear-late normalized weights w_t.
    """
    wb = load_workbook(value_xlsx_path, data_only=True)
    if value_sheet not in wb.sheetnames:
        raise RuntimeError(f"Sheet '{value_sheet}' not found in xlsx: {value_xlsx_path}")
    ws = wb[value_sheet]

    # Infer seq_len from first step block.
    # Row1: "Step 1", Row2: header, then data rows until blank first col.
    seq_len = 0
    row = 3
    while True:
        cell = ws.cell(row=row, column=1).value
        if cell is None or str(cell).strip() == "":
            break
        seq_len += 1
        row += 1
    if seq_len <= 0:
        raise RuntimeError(f"Cannot parse seq_len from xlsx: {value_xlsx_path}")

    # Parse all step blocks and delta-attention blocks.
    step_attn: List[np.ndarray] = []
    delta_attn: List[np.ndarray] = []

    max_row = ws.max_row
    row = 1
    while row <= max_row:
        title = ws.cell(row=row, column=1).value
        if isinstance(title, str) and title.startswith("Step ") and "Delta" not in title:
            # title row, header row, then data starts at row+2
            attn_t, _ = _parse_step_block_from_xlsx(ws, row + 2, seq_len)
            step_attn.append(attn_t)
            row += 2 + seq_len + 1
            continue
        if isinstance(title, str) and title.startswith("Step ") and "Delta" in title:
            d_attn_t = _parse_delta_attn_block_from_xlsx(ws, row + 2, seq_len)
            delta_attn.append(d_attn_t)
            row += 2 + seq_len + 1
            continue
        row += 1

    if len(step_attn) == 0:
        raise RuntimeError(f"No step blocks found in xlsx: {value_xlsx_path}")

    K = len(step_attn)
    L = seq_len

    # Build per-step DeltaA aligned with step index:
    # step_1 has no previous delta -> zeros.
    deltaA_by_step: List[np.ndarray] = [np.zeros((L, L), dtype=np.float64)]
    for t in range(1, K):
        if (t - 1) < len(delta_attn):
            deltaA_by_step.append(delta_attn[t - 1])
        else:
            deltaA_by_step.append(np.zeros((L, L), dtype=np.float64))

    # Linear-late weights, normalized.
    weights = np.arange(1, K + 1, dtype=np.float64)
    weights = weights / weights.sum()

    step_scores: List[np.ndarray] = []
    step_scores_no_local: List[np.ndarray] = []
    sym_edge_scores: List[np.ndarray] = []
    for t in range(K):
        A = step_attn[t]
        dA = deltaA_by_step[t]

        sym_attn = 0.5 * (A + A.T)
        sym_edge_scores.append(sym_attn)
        stable_gate = np.exp(-float(stability_lambda) * np.abs(dA))
        S_t_no_local = sym_attn * stable_gate
        step_scores_no_local.append(S_t_no_local)

        idx = np.arange(L, dtype=np.float64)
        local_dist = np.abs(idx[:, None] - idx[None, :])
        local_gate = np.exp(-float(local_beta) * local_dist)
        S_t = S_t_no_local * local_gate
        step_scores.append(S_t)

    side_mat_no_local = np.zeros((L, L), dtype=np.float64)
    for w, S_t0 in zip(weights, step_scores_no_local):
        side_mat_no_local += w * S_t0

    idx = np.arange(L, dtype=np.float64)
    local_dist = np.abs(idx[:, None] - idx[None, :])
    local_gate = np.exp(-float(local_beta) * local_dist)
    side_mat = side_mat_no_local * local_gate

    sym_edge_agg = np.zeros((L, L), dtype=np.float64)
    for w, sym_t in zip(weights, sym_edge_scores):
        sym_edge_agg += w * sym_t

    os.makedirs(os.path.dirname(side_xlsx_path) or ".", exist_ok=True)
    out_wb = Workbook()
    agg_ws = out_wb.active
    agg_ws.title = "side_aggregate"

    # Header
    agg_ws.append(["seq_idx"] + [f"j_{j + 1}" for j in range(L)])
    for i in range(L):
        agg_ws.append([i + 1] + side_mat[i, :].astype(float).tolist())

    # Save per-step S_t matrices for detailed inspection.
    for t, S_t in enumerate(step_scores, start=1):
        step_ws = out_wb.create_sheet(f"step_{t}")
        step_ws.append(["seq_idx"] + [f"j_{j + 1}" for j in range(L)])
        for i in range(L):
            step_ws.append([i + 1] + S_t[i, :].astype(float).tolist())

    # Symmetric edge score matrices: Sym_t = (A_t + A_t^T)/2
    sym_ws = out_wb.create_sheet("symmetric_edge_scores")
    sym_ws.append(["Symmetric Edge Score Aggregate", ""])
    sym_ws.append(["seq_idx"] + [f"j_{j + 1}" for j in range(L)])
    for i in range(L):
        sym_ws.append([i + 1] + sym_edge_agg[i, :].astype(float).tolist())
    sym_ws.append([])
    for t, sym_t in enumerate(sym_edge_scores, start=1):
        sym_ws.append([f"Step {t}", ""])
        sym_ws.append(["seq_idx"] + [f"j_{j + 1}" for j in range(L)])
        for i in range(L):
            sym_ws.append([i + 1] + sym_t[i, :].astype(float).tolist())
        sym_ws.append([])

    info_ws = out_wb.create_sheet("meta")
    info_ws.append(["formula", "M = (sum_t w_t * A_sym_t * exp(-lambda*|DeltaA_t|)) * exp(-beta*|i-j|)"])
    info_ws.append(["stability_lambda", float(stability_lambda)])
    info_ws.append(["local_beta", float(local_beta)])
    info_ws.append(["value_sheet", value_sheet])
    info_ws.append(["num_steps", int(K)])
    info_ws.append(["seq_len", int(L)])
    info_ws.append(["weights"] + weights.astype(float).tolist())

    out_wb.save(side_xlsx_path)
    return side_xlsx_path


class LLaDAValueRecorder:
    """Record value statistics from the first N denoising steps."""

    def __init__(self, model_path: str, device: str = "cuda"):
        self.device = device
        self.tokenizer = AutoTokenizer.from_pretrained(model_path, trust_remote_code=True)
        self.model = AutoModel.from_pretrained(
            model_path,
            torch_dtype=torch.bfloat16,
            trust_remote_code=True,
        ).to(device).eval()

        if hasattr(self.model, "transformer"):
            self.base_model = self.model
        elif hasattr(self.model, "model"):
            self.base_model = self.model.model
        else:
            raise AttributeError("Cannot find transformer in model")

        if hasattr(self.base_model, "transformer") and hasattr(self.base_model.transformer, "blocks"):
            self.blocks = self.base_model.transformer.blocks
        elif hasattr(self.base_model, "blocks"):
            self.blocks = self.base_model.blocks
        else:
            raise AttributeError("Cannot find transformer blocks")

        self.num_layers = len(self.blocks)
        self.mask_id = 126336

        config = self.base_model.config
        self.num_heads = config.n_heads
        self.d_model = config.d_model
        self.head_dim = self.d_model // self.num_heads
        self.num_kv_heads = getattr(config, "effective_n_kv_heads", self.num_heads)

    @staticmethod
    def _repeat_kv(hidden_states: torch.Tensor, n_rep: int) -> torch.Tensor:
        if n_rep == 1:
            return hidden_states
        batch, num_key_value_heads, seq_len, head_dim = hidden_states.shape
        hidden_states = hidden_states[:, :, None, :, :].expand(
            batch, num_key_value_heads, n_rep, seq_len, head_dim
        )
        return hidden_states.reshape(batch, num_key_value_heads * n_rep, seq_len, head_dim)

    def build_attention_bias(
        self,
        attention_mask: Optional[torch.Tensor],
        seq_len: int,
        dtype: torch.dtype,
        device: torch.device,
    ) -> Optional[torch.Tensor]:
        if attention_mask is None:
            return None
        mask = attention_mask[:, None, None, :].to(dtype=torch.float32, device=device)
        mask = (1.0 - mask) * torch.finfo(torch.float32).min
        mask = mask.expand(attention_mask.shape[0], 1, seq_len, seq_len)
        return mask.to(dtype=dtype)

    def extract_qkv_and_apply_rope(
        self, block, hidden_states: torch.Tensor
    ) -> Tuple[torch.Tensor, torch.Tensor, torch.Tensor]:
        dtype = hidden_states.dtype
        x_normed = block.attn_norm(hidden_states)

        q = block.q_proj(x_normed)
        k = block.k_proj(x_normed)
        v = block.v_proj(x_normed)

        batch, seq_len, width = q.size()

        if getattr(block, "q_norm", None) is not None and getattr(block, "k_norm", None) is not None:
            q = block.q_norm(q).to(dtype=dtype)
            k = block.k_norm(k).to(dtype=dtype)

        q = q.view(batch, seq_len, self.num_heads, width // self.num_heads).transpose(1, 2)
        k = k.view(batch, seq_len, self.num_kv_heads, width // self.num_heads).transpose(1, 2)
        v = v.view(batch, seq_len, self.num_kv_heads, width // self.num_heads).transpose(1, 2)

        if hasattr(block, "rotary_emb"):
            q, k = block.rotary_emb(q, k)

        return q, k, v

    def compute_attention_weights(
        self,
        q: torch.Tensor,
        k: torch.Tensor,
        attention_mask: Optional[torch.Tensor] = None,
    ) -> torch.Tensor:
        attn_weights = torch.matmul(q, k.transpose(-2, -1)) / math.sqrt(self.head_dim)

        if attention_mask is not None:
            seq_len = q.shape[-2]
            attn_bias = self.build_attention_bias(
                attention_mask=attention_mask,
                seq_len=seq_len,
                dtype=attn_weights.dtype,
                device=attn_weights.device,
            )
            attn_weights = attn_weights + attn_bias

        return torch.softmax(attn_weights, dim=-1, dtype=torch.float32).to(q.dtype)

    def compute_step_avg_attention(
        self,
        all_hidden_states,
        layers_to_extract: List[int],
        attention_mask: Optional[torch.Tensor] = None,
    ) -> torch.Tensor:
        attn_list = []
        for layer_idx in layers_to_extract:
            hidden = all_hidden_states[layer_idx]
            block = self.blocks[layer_idx]

            q, k, _ = self.extract_qkv_and_apply_rope(block, hidden)
            if self.num_kv_heads != self.num_heads:
                n_rep = self.num_heads // self.num_kv_heads
                k = self._repeat_kv(k, n_rep)

            attn = self.compute_attention_weights(q, k, attention_mask=attention_mask)
            attn_list.append(attn.mean(dim=1).squeeze(0))

        return torch.stack(attn_list, dim=0).mean(dim=0)

    @staticmethod
    def _decode_token_str(tokenizer, token_id: int) -> str:
        if int(token_id) < 0:
            return ""
        text = tokenizer.decode([int(token_id)], skip_special_tokens=False)
        return text.replace("\x00", "")

    def _decode_with_mask(self, token_ids: List[int]) -> str:
        parts: List[str] = []
        buf: List[int] = []
        for tid in token_ids:
            if int(tid) == self.mask_id:
                if buf:
                    txt = self.tokenizer.decode(buf, skip_special_tokens=False).replace("\x00", "")
                    parts.append(txt)
                    buf = []
                parts.append("<MASK>")
            else:
                buf.append(int(tid))
        if buf:
            txt = self.tokenizer.decode(buf, skip_special_tokens=False).replace("\x00", "")
            parts.append(txt)
        return "".join(parts)

    @torch.no_grad()
    def generate_and_record_values(
        self,
        prompt: torch.Tensor,
        steps: int = 128,
        gen_length: int = 512,
        block_length: int = 512,
        temperature: float = 0.0,
        cfg_scale: float = 0.0,
        capture_steps: int = 5,
        capture_start_step: int = 1,
        capture_end_step: Optional[int] = None,
        layers_to_extract: Optional[List[int]] = None,
        save_intermediate: bool = False,
        output_file: str = "/home/yzx/LLaDA/yzx_test/value_outputs/denoise_value_log.txt",
    ) -> Dict[str, object]:
        prompt_len = int(prompt.shape[1])
        x = torch.full(
            (1, prompt_len + gen_length),
            self.mask_id,
            dtype=torch.long,
            device=self.device,
        )
        x[:, :prompt_len] = prompt.clone()
        prompt_index = x != self.mask_id
        attention_mask = torch.ones_like(x, dtype=torch.long, device=self.device)

        assert gen_length % block_length == 0
        num_blocks = gen_length // block_length
        assert steps % num_blocks == 0
        steps_per_block = steps // num_blocks

        if layers_to_extract is None:
            layers_to_extract = list(range(max(0, self.num_layers - 2), self.num_layers))

        if capture_start_step < 1:
            raise ValueError("capture_start_step must be >= 1")
        if capture_end_step is not None and capture_end_step < capture_start_step:
            raise ValueError("capture_end_step must be >= capture_start_step")
        if capture_end_step is None:
            capture_end_step = capture_start_step + capture_steps - 1

        if save_intermediate:
            output_dir = os.path.dirname(output_file)
            if output_dir:
                os.makedirs(output_dir, exist_ok=True)
            with open(output_file, "w", encoding="utf-8") as f:
                f.write(
                    f"steps={steps}, gen_length={gen_length}, block_length={block_length}, "
                    f"capture_steps={capture_steps}, "
                    f"capture_start_step={capture_start_step}, capture_end_step={capture_end_step}\n"
                )
                f.write("=" * 60 + "\n")

        captured_records: List[Dict[str, object]] = []
        global_step = 0

        for num_block in range(num_blocks):
            block_start = prompt_len + num_block * block_length
            block_end = prompt_len + (num_block + 1) * block_length
            block_mask_index = x[:, block_start:block_end] == self.mask_id
            num_transfer_tokens = get_num_transfer_tokens(block_mask_index, steps_per_block)

            for i in range(steps_per_block):
                mask_index = x == self.mask_id
                if not mask_index[:, block_start:block_end].any():
                    global_step += 1
                    continue

                if cfg_scale > 0.0:
                    un_x = x.clone()
                    un_x[prompt_index] = self.mask_id
                    x_ = torch.cat([x, un_x], dim=0)
                    attention_mask_ = torch.cat([attention_mask, attention_mask], dim=0)

                    outputs = self.base_model(x_, attention_mask=attention_mask_, output_hidden_states=True)
                    logits = outputs.logits
                    logits, un_logits = torch.chunk(logits, 2, dim=0)
                    logits = un_logits + (cfg_scale + 1) * (logits - un_logits)
                    all_hidden_states = tuple(h[:1] for h in outputs.hidden_states)
                else:
                    outputs = self.base_model(x, attention_mask=attention_mask, output_hidden_states=True)
                    logits = outputs.logits
                    all_hidden_states = outputs.hidden_states

                avg_attn = self.compute_step_avg_attention(
                    all_hidden_states=all_hidden_states,
                    layers_to_extract=layers_to_extract,
                    attention_mask=attention_mask,
                )

                probs = F.softmax(logits, dim=-1)
                top_confidence, top_token_ids = torch.max(probs, dim=-1)

                logits_with_noise = add_gumbel_noise(logits, temperature=temperature)
                x0 = torch.argmax(logits_with_noise, dim=-1)
                x0_p = torch.squeeze(torch.gather(probs, dim=-1, index=torch.unsqueeze(x0, -1)), -1)

                x0_p[:, block_end:] = -np.inf
                x0 = torch.where(mask_index, x0, x)
                confidence = torch.where(mask_index, x0_p, -np.inf)
                transfer_index = torch.zeros_like(x0, dtype=torch.bool, device=x0.device)
                for j in range(x0.shape[0]):
                    k_this_step = int(num_transfer_tokens[j, i].item())
                    if k_this_step <= 0:
                        continue
                    _, select_index = torch.topk(confidence[j], k=k_this_step)
                    transfer_index[j, select_index] = True

                selected_positions = torch.where(transfer_index[0])[0]
                remain_ratio = count_remaining_mask_ratio(x, prompt_len, self.mask_id)

                current_step_1based = global_step + 1
                capture_this_step = (
                    capture_start_step <= current_step_1based <= capture_end_step
                )
                if capture_this_step:
                    record = {
                        "global_step": int(global_step),
                        "global_step_1based": int(current_step_1based),
                        "block_index": int(num_block),
                        "step_in_block": int(i),
                        "mask_response_before_update": mask_index[0, prompt_len:].detach().cpu().numpy().astype(bool),
                        "sequence_before_update": x[0].detach().cpu().numpy().astype(np.int64),
                        "avg_attention": avg_attn.detach().float().cpu().numpy(),
                        "top_confidence": top_confidence[0].detach().float().cpu().numpy(),
                        "top_token_ids": top_token_ids[0].detach().cpu().numpy().astype(np.int64),
                        "selected_positions": selected_positions.detach().cpu().numpy().astype(np.int64),
                        "selected_token_ids": x0[0, selected_positions].detach().cpu().numpy().astype(np.int64),
                    }

                x[transfer_index] = x0[transfer_index]
                if capture_this_step:
                    record["sequence_after_update"] = x[0].detach().cpu().numpy().astype(np.int64)
                    captured_records.append(record)

                if save_intermediate:
                    response_ids = x[0, prompt_len:].detach().tolist()
                    intermediate_text = self._decode_with_mask(response_ids)
                    with open(output_file, "a", encoding="utf-8") as f:
                        f.write(
                            f"GlobalStep {global_step + 1}/{steps} "
                            f"Block {num_block + 1}/{num_blocks} "
                            f"StepInBlock {i + 1}/{steps_per_block} "
                            f"selected={int(selected_positions.numel())} "
                            f"remain_ratio={remain_ratio:.4f}\n"
                        )
                        f.write(intermediate_text + "\n")
                        f.write("-" * 60 + "\n")

                global_step += 1

        if len(captured_records) == 0:
            raise RuntimeError("No denoising step was captured. Please check generation settings.")

        attention_scores = np.stack([r["avg_attention"] for r in captured_records], axis=0)
        top_confidences = np.stack([r["top_confidence"] for r in captured_records], axis=0)
        top_token_ids = np.stack([r["top_token_ids"] for r in captured_records], axis=0)
        seq_before = np.stack([r["sequence_before_update"] for r in captured_records], axis=0)
        seq_after = np.stack([r["sequence_after_update"] for r in captured_records], axis=0)
        mask_response = np.stack([r["mask_response_before_update"] for r in captured_records], axis=0)

        # Keep only response-region values.
        attention_scores = attention_scores[:, prompt_len:, prompt_len:]
        top_confidences = top_confidences[:, prompt_len:]
        top_token_ids = top_token_ids[:, prompt_len:]
        seq_before = seq_before[:, prompt_len:]
        seq_after = seq_after[:, prompt_len:]

        attention_scores_all_all = attention_scores.copy()
        top_confidences_all = top_confidences.copy()
        top_token_ids_all = top_token_ids.copy()

        mask_row = mask_response[:, :, None]
        mask_pair = np.logical_and(mask_response[:, :, None], mask_response[:, None, :])
        attention_scores_mask_all = np.where(mask_row, attention_scores_all_all, 0.0)
        attention_scores_mask_mask = np.where(mask_pair, attention_scores_all_all, 0.0)
        top_confidences_mask = np.where(mask_response, top_confidences_all, 0.0)
        top_token_ids_mask = np.where(mask_response, top_token_ids_all, -1)

        if len(captured_records) > 1:
            attention_deltas_all_all = attention_scores_all_all[1:] - attention_scores_all_all[:-1]
            attention_deltas_mask_all = attention_scores_mask_all[1:] - attention_scores_mask_all[:-1]
            attention_deltas_mask_mask = attention_scores_mask_mask[1:] - attention_scores_mask_mask[:-1]
            confidence_deltas_all = top_confidences_all[1:] - top_confidences_all[:-1]
            confidence_deltas_mask = top_confidences_mask[1:] - top_confidences_mask[:-1]

            mask_both_nodes = np.logical_and(mask_response[1:], mask_response[:-1])
            mask_both_rows = mask_both_nodes[:, :, None]
            mask_both_pairs = np.logical_and(mask_both_nodes[:, :, None], mask_both_nodes[:, None, :])
            attention_deltas_mask_all = np.where(mask_both_rows, attention_deltas_mask_all, 0.0)
            attention_deltas_mask_mask = np.where(mask_both_pairs, attention_deltas_mask_mask, 0.0)
            confidence_deltas_mask = np.where(mask_both_nodes, confidence_deltas_mask, 0.0)
        else:
            empty_attn = np.empty((0, *attention_scores_all_all.shape[1:]), dtype=np.float32)
            empty_conf = np.empty((0, top_confidences_all.shape[1]), dtype=np.float32)
            attention_deltas_all_all = empty_attn
            attention_deltas_mask_all = empty_attn.copy()
            attention_deltas_mask_mask = empty_attn.copy()
            confidence_deltas_all = empty_conf
            confidence_deltas_mask = empty_conf.copy()

        decoded_top_tokens = [
            [self._decode_token_str(self.tokenizer, tid) for tid in step_ids]
            for step_ids in top_token_ids_all
        ]

        meta = {
            "num_captured_steps": len(captured_records),
            "capture_steps_target": capture_steps,
            "capture_start_step": int(capture_start_step),
            "capture_end_step": int(capture_end_step),
            "prompt_length": prompt_len,
            "layers_to_extract": layers_to_extract,
            "global_steps": [r["global_step"] for r in captured_records],
            "global_steps_1based": [r["global_step_1based"] for r in captured_records],
            "block_indices": [r["block_index"] for r in captured_records],
            "steps_in_block": [r["step_in_block"] for r in captured_records],
            "selected_positions": [r["selected_positions"].tolist() for r in captured_records],
            "selected_token_ids": [r["selected_token_ids"].tolist() for r in captured_records],
            "selected_tokens": [
                [self._decode_token_str(self.tokenizer, t) for t in r["selected_token_ids"]]
                for r in captured_records
            ],
            "top_token_text_per_step": decoded_top_tokens,
            "mask_response_before_update": mask_response.astype(np.int8).tolist(),
        }

        return {
            "final_response_sequence": x[0, prompt_len:].detach().cpu().numpy().astype(np.int64),
            "attention_scores_all_all": attention_scores_all_all.astype(np.float32),
            "attention_scores_mask_all": attention_scores_mask_all.astype(np.float32),
            "attention_scores_mask_mask": attention_scores_mask_mask.astype(np.float32),
            "top_confidences_all": top_confidences_all.astype(np.float32),
            "top_confidences_mask": top_confidences_mask.astype(np.float32),
            "top_token_ids_all": top_token_ids_all.astype(np.int64),
            "top_token_ids_mask": top_token_ids_mask.astype(np.int64),
            "attention_deltas_all_all": attention_deltas_all_all.astype(np.float32),
            "attention_deltas_mask_all": attention_deltas_mask_all.astype(np.float32),
            "attention_deltas_mask_mask": attention_deltas_mask_mask.astype(np.float32),
            "confidence_deltas_all": confidence_deltas_all.astype(np.float32),
            "confidence_deltas_mask": confidence_deltas_mask.astype(np.float32),
            "sequence_before_update": seq_before.astype(np.int64),
            "sequence_after_update": seq_after.astype(np.int64),
            "meta": meta,
        }

    @staticmethod
    def save_values(record: Dict[str, object], output_prefix: str) -> str:
        output_dir = os.path.dirname(output_prefix)
        if output_dir:
            os.makedirs(output_dir, exist_ok=True)

        xlsx_path = f"{output_prefix}.xlsx"

        meta = record["meta"]
        top_token_text_per_step = meta.get("top_token_text_per_step", [])
        prompt_len = int(meta.get("prompt_length", 0))

        wb = Workbook()
        default_ws = wb.active
        wb.remove(default_ws)

        _write_value_sheet(
            wb=wb,
            sheet_name="denoise_values_all_all",
            attention_scores=record["attention_scores_all_all"],
            top_confidences=record["top_confidences_all"],
            top_token_ids=record["top_token_ids_all"],
            attention_deltas=record["attention_deltas_all_all"],
            confidence_deltas=record["confidence_deltas_all"],
            prompt_len=prompt_len,
            top_token_text_per_step=top_token_text_per_step,
        )
        _write_value_sheet(
            wb=wb,
            sheet_name="denoise_values_mask_all",
            attention_scores=record["attention_scores_mask_all"],
            top_confidences=record["top_confidences_mask"],
            top_token_ids=record["top_token_ids_mask"],
            attention_deltas=record["attention_deltas_mask_all"],
            confidence_deltas=record["confidence_deltas_mask"],
            prompt_len=prompt_len,
            top_token_text_per_step=top_token_text_per_step,
        )
        _write_value_sheet(
            wb=wb,
            sheet_name="denoise_values_mask_mask",
            attention_scores=record["attention_scores_mask_mask"],
            top_confidences=record["top_confidences_mask"],
            top_token_ids=record["top_token_ids_mask"],
            attention_deltas=record["attention_deltas_mask_mask"],
            confidence_deltas=record["confidence_deltas_mask"],
            prompt_len=prompt_len,
            top_token_text_per_step=top_token_text_per_step,
        )
        wb["denoise_values_mask_all"].title = "denoise_values"

        meta_ws = wb.create_sheet("meta")
        meta_ws.append(["default_value_sheet", "denoise_values"])
        meta_ws.append(["all_all_sheet", "denoise_values_all_all"])
        meta_ws.append(["mask_all_sheet", "denoise_values"])
        meta_ws.append(["mask_mask_sheet", "denoise_values_mask_mask"])
        meta_ws.append(["capture_start_step", int(meta.get("capture_start_step", 1))])
        meta_ws.append(["capture_end_step", int(meta.get("capture_end_step", 1))])
        meta_ws.append(["num_captured_steps", int(meta.get("num_captured_steps", 0))])

        wb.save(xlsx_path)
        return xlsx_path


def main():
    prompt =""" 
        [
      {
        "role": "system",
        "content": "You are a helpful assistant to make travel plans for Bob.\n\nEXTERNAL RESOURCES:\n1. A database containing information about train tickets, attractions, and city transportation.\n2. A python notebook to execute python code for numerical operations and planning. \n\nTASK DESCRIPTION\nYou need to make a travel plan based on the given requirements, taking into account transportation between cities and daily schedules.\nThe final travel plan may include or be part of the following:\n1.go_to_place(origin:str,destination:str,departure_time,arrival_time): go to destination from origin. The origin and destination should be the name of a hotel or a spot instead of a city.\n2.visit(place:str,begin_time,end_time): visit somewhere from begin_time to end_time. The time should be expressed as \"%Y-%m-%d %H:%M\", e.g. 2023-07-02 16:00. You have to go somewhere before you can visit it.\n3.go_to_city(origin_city:str,destination_city:str,departure_time,arrival_time,ticket_number): go to destination city from origin city, using the ticket with the ticket_number(you can know the ticket number from the database).\n4.stay_in(city:str,begin_time,end_time): stay in somewhere from begin_time to end_time. The time should be expressed as \"%Y-%m-%d %H:%M\". Only when Bob is in some city can he visit it.\ne.g. \n<plan>go_to_place(\"Beijing Railway Hotel\",\"The Great Wall\",\"2023-07-02 7:00\",\"2023-07-02 8:05\")</plan>, <plan>visit(\"The Great Wall\",\"2023-07-02 8:05\",\"2023-07-05 17:00\")</plan>,<plan>go_to_city(\"Shanghai\",\"Beijing\",\"2023-07-02 16:00\",\"2023-07-02 22:30\",\"D1111\")</plan>, <plan>stay_in(\"Beijing\",\"2023-07-02 22:30\",\"2023-07-05 8:00\")</plan>\nYour ultimate goal is to give these plans, there is no need to do anything extra.\n\n--- Your Workflow ---\n1. You will first be given a task.\n2. You should understand the task and devise a plan to complete the task. This plan will contain a series of subtasks that need to be completed.\n\nPLAN AND SUBTASK:\nIf the task cannot be easily solved directly or requires the use of external resources, please assign it to another agent to complete (such as \"find the cheapest train from Beijing to Shanghai in 2023-7-1\"), otherwise you can complete it yourself. You may need to wait for results from other agents before proceeding to the next step of the task. If you need help from other agents, please clearly describe the task objectives, background, and precautions of the subtask. \nA subtask-structure has the following json component and surrounded with <subtask></subtask> as follows:\n<subtask>{\n\"subtask_name\": string, name of the subtask\n\"goal\": string, the main purpose of the subtask, and what will you do to reach this goal?\n\"criticism\": string, what potential problems may the current subtask and goal have?\n\"milestones\": list[string]. what milestones should be achieved to ensure the subtask is done? Make it detailed and specific.\n\"result_format\": optional, what the result should be.}</subtask>\n\n"
      },
      {
        "role": "system",
        "name": "example_user",
        "content": "Task Requirements: Bob is in Shanghai and going to travel in several cities, please make a ticket purchase plan and travel sequence for him.The demands are as follows:\n1. visit ['Beijing']. The order doesn't matter and he needs to return to Shanghai finally.\n2. He is free to travel from 2023.7.1 to 2023.7.20. The budget for transportation is 1000.0 CNY.\n3. Play at least 3 days in Beijing.\n4. If you arrive in a city before 12:00 noon, that day can be counted as a day of play. If it's past 12 o'clock, it doesn't count as a day.\n5. On the basis of completing the above conditions (especially the budget), spend as little time as possible.\n"
      },
      {
        "role": "system",
        "name": "example_assistant",
        "content": "Based on the requirements, we can know that Bob need to go to Beijing from Shanghai, stay in Beijing for 3 days and then go to Shanghai from Beijing.\nGiven the task, the first step is to find available train tickets that fit Bob's schedule and budget.\n<subtask>\n{\n\"subtask_name\": \"find_available_train_tickets\",\n\"goal\": \"Find train tickets from Shanghai to Beijing and back to Shanghai that fit within the travel dates, budget, and allow for at least 3 full days of play in Beijing. If the arrival is before 12:00 noon, it counts as a day of play.\",\n\"criticism\": \"Must ensure that the total cost of the round trip tickets does not exceed the budget of 1000.0 CNY and that the timings allow for at least 3 full days in Beijing for visit. So you need to allow time between train rides(arrival in a city and departure from the city). For each ticket, you must give me the ticket number, origin, destination, departure time, arrival time and the price.\",\n\"milestones\": [\"Identify a suitable train from Shanghai to Beijing.\", \"Identify a return train from Beijing to Shanghai ensuring at least 3 days in Beijing before departing.\", \"Ensure the total cost of both tickets is within the budget of 1000.0 CNY.\"]\n}\n</subtask>\nThen we can get the final plan consists of go_to_city and stay_in.\n<subtask>\n{\n\"subtask_name\": \"get the final plan\",\n\"goal\": \"Formulate a travel plan for Bob's trip from Shanghai to Beijing and back, ensuring it fits within his budget and time constraints, including at least 3 full days in Beijing.\",\n\"criticism\": \"The plan must be concise, focusing on efficient travel and stay arrangements while adhering to the budget and time constraints.\",\n\"milestones\": [\"Include suitable train journeys within the budget.\",\"Plan at least 3 full days in Beijing.\",\"Ensure the overall plan fits within the specified dates and budget.\"],\n\"result_format\": \"A schedule consisting with multiple <plan>go_to_place(...)</plan> and <plan>visit(...)</plan>.    1.go_to_place(origin:str,destination:str,departure_time,arrival_time): go to destination from origin.     2.visit(place:str,begin_time,end_time): visit somewhere from begin_time to end_time. The time should be expressed as %Y-%m-%d %H:%M, e.g. 2023-07-02 16:00.\"\n}\n</subtask>\n"
      },
      {
        "role": "user",
        "content": "Task Requirements:\nBob is in Beijing and going to travel in several cities, please make a ticket purchase plan and travel sequence for him.The demands are as follows:\n1. visit ['Chengdu']. The order doesn't matter and he needs to return to Beijing finally.\n2. He is free to travel from 2023.7.1 to 2023.7.20. The budget for transportation is 1800.0 CNY.\n3. Play at least 1 day in Chengdu.\n4. Stay in any city for a minimum of 24 hours to count as one day.\n5. On the basis of completing the above conditions (especially the budget), spend as little time as possible.\nCome up with an abstract plan to perform this task in a couple of steps. Give me the subtasks between <subtask> and </subtask>."
      }
    ]
    """
    parser = argparse.ArgumentParser(description="Record first denoising-step values from LLaDA generation.")
    parser.add_argument("--model-path", type=str, default="/data/labshare/Param/llada/")
    parser.add_argument("--device", type=str, default="cuda")
    parser.add_argument("--prompt", type=str, default=prompt)
    parser.add_argument("--gen-length", type=int, default=512)
    parser.add_argument("--steps", type=int, default=128)
    parser.add_argument("--block-length", type=int, default=512)
    parser.add_argument("--capture-steps", type=int, default=5)
    parser.add_argument(
        "--capture-start-step",
        type=int,
        default=1,
        help="1-based global denoising step to start capturing from",
    )
    parser.add_argument(
        "--capture-end-step",
        type=int,
        default=None,
        help="1-based global denoising step to stop capturing at, inclusive",
    )
    parser.add_argument("--temperature", type=float, default=0.0)
    parser.add_argument("--cfg-scale", type=float, default=0.0)
    parser.add_argument("--save-intermediate", action="store_true")
    parser.add_argument("--stability-lambda", type=float, default=1.0)
    parser.add_argument("--local-beta", type=float, default=0.05)
    parser.add_argument(
        "--value-sheet",
        type=str,
        default="denoise_values",
        choices=["denoise_values", "denoise_values_all_all", "denoise_values_mask_mask"],
        help="Which value sheet to use when computing side.xlsx; denoise_values defaults to mask-all",
    )
    parser.add_argument(
        "--side-output",
        type=str,
        default="/home/yzx/LLaDA/yzx_test/value_outputs/side.xlsx",
    )
    
    parser.add_argument(
        "--output-file",
        type=str,
        default="/home/yzx/LLaDA/yzx_test/denoise_value_log.txt",
    )
    parser.add_argument(
        "--output-prefix",
        type=str,
        default="/home/yzx/LLaDA/yzx_test/value_outputs/denoise_value",
    )
    parser.set_defaults(save_intermediate=True)
    args = parser.parse_args()

    recorder = LLaDAValueRecorder(model_path=args.model_path, device=args.device)

    input_text = f"Question: {args.prompt}\nAnswer:"
    inputs = recorder.tokenizer(input_text, return_tensors="pt").to(args.device)

    record = recorder.generate_and_record_values(
        prompt=inputs.input_ids,
        steps=args.steps,
        gen_length=args.gen_length,
        block_length=args.block_length,
        temperature=args.temperature,
        cfg_scale=args.cfg_scale,
        capture_steps=args.capture_steps,
        capture_start_step=args.capture_start_step,
        capture_end_step=args.capture_end_step,
        save_intermediate=args.save_intermediate,
        output_file=args.output_file,
    )

    xlsx_path = recorder.save_values(record, args.output_prefix)
    side_xlsx_path = compute_side_xlsx_from_value_xlsx(
        value_xlsx_path=xlsx_path,
        side_xlsx_path=args.side_output,
        stability_lambda=args.stability_lambda,
        local_beta=args.local_beta,
        value_sheet=args.value_sheet,
    )
    print(f"Saved values to xlsx: {xlsx_path}")
    print(f"Saved side matrix to xlsx: {side_xlsx_path}")
    if args.save_intermediate:
        print(f"Saved intermediate log to: {args.output_file}")
    print(
        "Final generated response: "
        f"{recorder.tokenizer.decode(record['final_response_sequence'], skip_special_tokens=True)}"
    )


if __name__ == "__main__":
    main()
