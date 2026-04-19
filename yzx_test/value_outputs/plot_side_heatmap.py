import argparse
import os
from typing import List, Optional

import numpy as np
from openpyxl import load_workbook


def _read_matrix_from_sheet(sheet) -> np.ndarray:
    # Expected format:
    # row1: ["seq_idx", "j_1", ...]
    # row2..: [i, v11, v12, ...]
    max_row = sheet.max_row
    max_col = sheet.max_column
    if max_row < 2 or max_col < 2:
        raise ValueError(f"Sheet '{sheet.title}' has no matrix data.")

    values: List[List[float]] = []
    for r in range(2, max_row + 1):
        row_vals = []
        for c in range(2, max_col + 1):
            v = sheet.cell(row=r, column=c).value
            row_vals.append(float(v) if v is not None else 0.0)
        values.append(row_vals)
    return np.array(values, dtype=np.float32)


def plot_heatmap(
    matrix: np.ndarray,
    output_path: str,
    title: str,
    vmin: Optional[float] = None,
    vmax: Optional[float] = None,
):
    try:
        import matplotlib

        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        from matplotlib.colors import LinearSegmentedColormap
    except Exception as e:
        raise RuntimeError(
            "matplotlib is required for plotting. "
            "Please install it in your runtime environment."
        ) from e

    h, w = matrix.shape
    if vmin is None:
        vmin = float(np.nanmin(matrix))
    if vmax is None:
        vmax = float(np.nanmax(matrix))
    if vmax <= vmin:
        vmax = vmin + 1e-6

    figsize = (max(6, w * 0.15), max(5, h * 0.15))
    plt.figure(figsize=figsize)
    white_green_blue = LinearSegmentedColormap.from_list(
        "white_green_blue",
        [
            (0.00, "#ffffff"),
            (0.35, "#dff3e3"),
            (0.60, "#6bcf8a"),
            (0.82, "#1d91c0"),
            (1.00, "#084081"),
        ],
    )
    im = plt.imshow(
        matrix,
        cmap=white_green_blue,
        aspect="auto",
        interpolation="nearest",
        vmin=vmin,
        vmax=vmax,
    )
    plt.colorbar(im, fraction=0.046, pad=0.04, label="Dependency Score")
    plt.title(title)
    plt.xlabel("Token j")
    plt.ylabel("Token i")
    plt.tight_layout()
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    plt.savefig(output_path, dpi=220, bbox_inches="tight")
    plt.close()


def main():
    parser = argparse.ArgumentParser(description="Plot dependency heatmap from side.xlsx")
    parser.add_argument(
        "--side-xlsx",
        type=str,
        default="/home/yzx/LLaDA/yzx_test/value_outputs/side.xlsx",
        help="Path to side.xlsx",
    )
    parser.add_argument(
        "--sheet",
        type=str,
        default="side_aggregate",
        help="Sheet to plot, e.g. side_aggregate or step_1",
    )
    parser.add_argument(
        "--output",
        type=str,
        default="/home/yzx/LLaDA/yzx_test/value_outputs/side_heatmap.png",
        help="Output image path",
    )
    parser.add_argument("--vmin", type=float, default=None, help="Heatmap min value; default is auto")
    parser.add_argument("--vmax", type=float, default=None, help="Heatmap max value; default is auto")
    args = parser.parse_args()

    wb = load_workbook(args.side_xlsx, data_only=True)
    if args.sheet not in wb.sheetnames:
        raise ValueError(f"Sheet '{args.sheet}' not found. Available: {wb.sheetnames}")

    matrix = _read_matrix_from_sheet(wb[args.sheet])
    plot_heatmap(
        matrix=matrix,
        output_path=args.output,
        title=f"{args.sheet} (range: auto)" if args.vmin is None and args.vmax is None else f"{args.sheet} (range: {args.vmin}~{args.vmax})",
        vmin=args.vmin,
        vmax=args.vmax,
    )
    print(f"Saved heatmap to: {args.output}")


if __name__ == "__main__":
    main()
