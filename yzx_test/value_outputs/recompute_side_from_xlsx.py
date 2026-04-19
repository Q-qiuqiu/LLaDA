import argparse
import json
import os
from typing import List, Tuple

import numpy as np
from openpyxl import Workbook, load_workbook


def _parse_step_block_from_xlsx(sheet, start_row: int, seq_len: int) -> np.ndarray:
    rows: List[List[float]] = []
    row = start_row
    for _ in range(seq_len):
        attn_json = sheet.cell(row=row, column=3).value
        rows.append(json.loads(attn_json) if isinstance(attn_json, str) else list(attn_json))
        row += 1
    return np.array(rows, dtype=np.float64)


def _load_value_xlsx(value_xlsx_path: str, value_sheet: str) -> Tuple[List[np.ndarray], List[np.ndarray], int]:
    wb = load_workbook(value_xlsx_path, data_only=True)
    if value_sheet not in wb.sheetnames:
        raise RuntimeError(f"Sheet '{value_sheet}' not found in xlsx: {value_xlsx_path}")
    ws = wb[value_sheet]

    seq_len = 0
    row = 3
    while True:
        cell = ws.cell(row=row, column=1).value
        if cell is None or str(cell).strip() == "":
            break
        seq_len += 1
        row += 1
    if seq_len <= 0:
        raise RuntimeError(f"Cannot parse seq_len from xlsx: {value_xlsx_path}")

    step_attn: List[np.ndarray] = []
    delta_attn: List[np.ndarray] = []
    row = 1
    while row <= ws.max_row:
        title = ws.cell(row=row, column=1).value
        if isinstance(title, str) and title.startswith("Step ") and "Delta" not in title:
            step_attn.append(_parse_step_block_from_xlsx(ws, row + 2, seq_len))
            row += 2 + seq_len + 1
            continue
        if isinstance(title, str) and title.startswith("Step ") and "Delta" in title:
            delta_attn.append(_parse_step_block_from_xlsx(ws, row + 2, seq_len))
            row += 2 + seq_len + 1
            continue
        row += 1

    if not step_attn:
        raise RuntimeError(f"No step blocks found in xlsx: {value_xlsx_path}")
    return step_attn, delta_attn, seq_len


def compute_side_matrices(
    step_attn: List[np.ndarray],
    delta_attn: List[np.ndarray],
    stability_lambda: float,
    local_beta: float,
) -> Tuple[np.ndarray, List[np.ndarray], List[np.ndarray], np.ndarray]:
    num_steps = len(step_attn)
    seq_len = step_attn[0].shape[0]

    delta_by_step: List[np.ndarray] = [np.zeros((seq_len, seq_len), dtype=np.float64)]
    for step_idx in range(1, num_steps):
        if step_idx - 1 < len(delta_attn):
            delta_by_step.append(delta_attn[step_idx - 1])
        else:
            delta_by_step.append(np.zeros((seq_len, seq_len), dtype=np.float64))

    weights = np.arange(1, num_steps + 1, dtype=np.float64)
    weights /= weights.sum()

    idx = np.arange(seq_len, dtype=np.float64)
    local_dist = np.abs(idx[:, None] - idx[None, :])
    local_gate = np.exp(-float(local_beta) * local_dist)

    step_scores: List[np.ndarray] = []
    sym_edge_scores: List[np.ndarray] = []
    side_mat = np.zeros((seq_len, seq_len), dtype=np.float64)
    for weight, attn, delta in zip(weights, step_attn, delta_by_step):
        sym_attn = 0.5 * (attn + attn.T)
        sym_edge_scores.append(sym_attn)
        stable_gate = np.exp(-float(stability_lambda) * np.abs(delta))
        step_score = sym_attn * stable_gate * local_gate
        step_scores.append(step_score)
        side_mat += weight * step_score

    sym_edge_agg = np.zeros((seq_len, seq_len), dtype=np.float64)
    for weight, sym_t in zip(weights, sym_edge_scores):
        sym_edge_agg += weight * sym_t

    return side_mat, step_scores, sym_edge_scores, weights


def save_side_xlsx(
    output_path: str,
    side_mat: np.ndarray,
    step_scores: List[np.ndarray],
    sym_edge_scores: List[np.ndarray],
    weights: np.ndarray,
    stability_lambda: float,
    local_beta: float,
    value_sheet: str,
) -> None:
    seq_len = side_mat.shape[0]
    wb = Workbook()
    agg_ws = wb.active
    agg_ws.title = "side_aggregate"
    agg_ws.append(["seq_idx"] + [f"j_{j + 1}" for j in range(seq_len)])
    for i in range(seq_len):
        agg_ws.append([i + 1] + side_mat[i, :].astype(float).tolist())

    for step_idx, step_score in enumerate(step_scores, start=1):
        ws = wb.create_sheet(f"step_{step_idx}")
        ws.append(["seq_idx"] + [f"j_{j + 1}" for j in range(seq_len)])
        for i in range(seq_len):
            ws.append([i + 1] + step_score[i, :].astype(float).tolist())

    sym_ws = wb.create_sheet("symmetric_edge_scores")
    sym_ws.append(["Symmetric Edge Score Aggregate", ""])
    sym_ws.append(["seq_idx"] + [f"j_{j + 1}" for j in range(seq_len)])
    sym_agg = np.zeros((seq_len, seq_len), dtype=np.float64)
    for weight, sym_t in zip(weights, sym_edge_scores):
        sym_agg += weight * sym_t
    for i in range(seq_len):
        sym_ws.append([i + 1] + sym_agg[i, :].astype(float).tolist())
    sym_ws.append([])
    for step_idx, sym_t in enumerate(sym_edge_scores, start=1):
        sym_ws.append([f"Step {step_idx}", ""])
        sym_ws.append(["seq_idx"] + [f"j_{j + 1}" for j in range(seq_len)])
        for i in range(seq_len):
            sym_ws.append([i + 1] + sym_t[i, :].astype(float).tolist())
        sym_ws.append([])

    info_ws = wb.create_sheet("meta")
    info_ws.append(["formula", "M = (sum_t w_t * A_sym_t * exp(-lambda*|DeltaA_t|)) * exp(-beta*|i-j|)"])
    info_ws.append(["stability_lambda", float(stability_lambda)])
    info_ws.append(["local_beta", float(local_beta)])
    info_ws.append(["value_sheet", value_sheet])
    info_ws.append(["num_steps", int(len(step_scores))])
    info_ws.append(["seq_len", int(seq_len)])
    info_ws.append(["weights"] + weights.astype(float).tolist())

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)


def save_heatmap(matrix: np.ndarray, output_path: str, title: str) -> None:
    try:
        import matplotlib

        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        from matplotlib.colors import LinearSegmentedColormap
    except Exception as e:
        raise RuntimeError(
            "matplotlib is required for plotting. Please install it in your runtime environment."
        ) from e

    vmin = float(np.nanmin(matrix))
    vmax = float(np.nanmax(matrix))
    if vmax <= vmin:
        vmax = vmin + 1e-6

    h, w = matrix.shape
    figsize = (max(6, w * 0.15), max(5, h * 0.15))
    plt.figure(figsize=figsize)
    cmap = LinearSegmentedColormap.from_list(
        "white_green_blue",
        [
            (0.00, "#ffffff"),
            (0.35, "#dff3e3"),
            (0.60, "#6bcf8a"),
            (0.82, "#1d91c0"),
            (1.00, "#084081"),
        ],
    )
    im = plt.imshow(matrix, cmap=cmap, aspect="auto", interpolation="nearest", vmin=vmin, vmax=vmax)
    plt.colorbar(im, fraction=0.046, pad=0.04, label="Dependency Score")
    plt.title(title)
    plt.xlabel("Token j")
    plt.ylabel("Token i")
    plt.tight_layout()
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    plt.savefig(output_path, dpi=220, bbox_inches="tight")
    plt.close()


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Recompute side.xlsx and heatmap directly from denoise_value.xlsx without rerunning inference."
    )
    parser.add_argument(
        "--value-xlsx",
        type=str,
        default="/home/yzx/LLaDA/yzx_test/value_outputs/denoise_value.xlsx",
        help="Path to denoise_value.xlsx",
    )
    parser.add_argument(
        "--value-sheet",
        type=str,
        default="denoise_values_mask_mask",
        help="Which sheet in value xlsx to read; denoise_values defaults to mask-all",
    )
    parser.add_argument(
        "--side-output",
        type=str,
        default="/home/yzx/LLaDA/yzx_test/value_outputs/side.xlsx",
        help="Path to save side.xlsx",
    )
    parser.add_argument(
        "--heatmap-output",
        type=str,
        default="/home/yzx/LLaDA/yzx_test/value_outputs/side_heatmap.png",
        help="Path to save heatmap image",
    )
    parser.add_argument("--stability-lambda", type=float, default=1000.0)
    parser.add_argument("--local-beta", type=float, default=0.005)
    args = parser.parse_args()

    step_attn, delta_attn, _ = _load_value_xlsx(args.value_xlsx, args.value_sheet)
    side_mat, step_scores, sym_edge_scores, weights = compute_side_matrices(
        step_attn=step_attn,
        delta_attn=delta_attn,
        stability_lambda=args.stability_lambda,
        local_beta=args.local_beta,
    )
    save_side_xlsx(
        output_path=args.side_output,
        side_mat=side_mat,
        step_scores=step_scores,
        sym_edge_scores=sym_edge_scores,
        weights=weights,
        stability_lambda=args.stability_lambda,
        local_beta=args.local_beta,
        value_sheet=args.value_sheet,
    )
    save_heatmap(
        matrix=side_mat,
        output_path=args.heatmap_output,
        title=(
            f"side_aggregate ({args.value_sheet}, "
            f"lambda={args.stability_lambda}, beta={args.local_beta})"
        ),
    )
    print(f"Saved side matrix xlsx to: {args.side_output}")
    print(f"Saved heatmap to: {args.heatmap_output}")


if __name__ == "__main__":
    main()
