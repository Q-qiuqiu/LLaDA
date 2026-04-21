import argparse
import html
import json
import os
from dataclasses import dataclass
from typing import List, Optional, Sequence, Tuple

import numpy as np
from openpyxl import Workbook, load_workbook


@dataclass
class StepBlock:
    step_idx: int
    seq_indices: np.ndarray
    tokens: List[str]
    attention: np.ndarray
    confidence: np.ndarray


SPECIAL_PATTERNS = ("<|endoftext|>", "<|eot_id|>")
LINE_BREAK_TOKENS = {"\n", "\\n"}


def _parse_json_array(value, expected_len: Optional[int] = None) -> np.ndarray:
    if isinstance(value, str):
        arr = json.loads(value)
    elif isinstance(value, (list, tuple)):
        arr = list(value)
    else:
        raise RuntimeError(f"Unsupported cell value type for array parsing: {type(value)}")
    result = np.asarray(arr, dtype=np.float64)
    if expected_len is not None and result.shape[0] != expected_len:
        raise RuntimeError(f"Expected array length {expected_len}, got {result.shape[0]}")
    return result


def _infer_seq_len(sheet) -> int:
    seq_len = 0
    row = 3
    while True:
        cell = sheet.cell(row=row, column=1).value
        if cell is None or str(cell).strip() == "":
            break
        seq_len += 1
        row += 1
    if seq_len <= 0:
        raise RuntimeError("Failed to infer sequence length from sheet.")
    return seq_len


def load_step_blocks(value_xlsx_path: str, value_sheet: str) -> List[StepBlock]:
    wb = load_workbook(value_xlsx_path, data_only=True)
    if value_sheet not in wb.sheetnames:
        raise RuntimeError(f"Sheet '{value_sheet}' not found in xlsx: {value_xlsx_path}")
    ws = wb[value_sheet]
    seq_len = _infer_seq_len(ws)

    step_blocks: List[StepBlock] = []
    row = 1
    while row <= ws.max_row:
        title = ws.cell(row=row, column=1).value
        if isinstance(title, str) and title.startswith("Step "):
            if "Delta" in title:
                row += 2 + seq_len + 1
                continue
            parts = title.split()
            if len(parts) < 2 or not parts[1].isdigit():
                raise RuntimeError(f"Cannot parse step title: {title}")
            step_idx = int(parts[1])
            seq_indices: List[int] = []
            tokens: List[str] = []
            confidence: List[float] = []
            attn_rows: List[np.ndarray] = []
            for offset in range(seq_len):
                cur_row = row + 2 + offset
                seq_indices.append(int(ws.cell(row=cur_row, column=1).value))
                tokens.append(str(ws.cell(row=cur_row, column=2).value))
                attn_rows.append(
                    _parse_json_array(ws.cell(row=cur_row, column=3).value, expected_len=seq_len)
                )
                conf_cell = ws.cell(row=cur_row, column=4).value
                confidence.append(0.0 if conf_cell is None else float(conf_cell))
            step_blocks.append(
                StepBlock(
                    step_idx=step_idx,
                    seq_indices=np.asarray(seq_indices, dtype=np.int64),
                    tokens=tokens,
                    attention=np.vstack(attn_rows),
                    confidence=np.asarray(confidence, dtype=np.float64),
                )
            )
            row += 2 + seq_len + 1
            continue
        row += 1
    if not step_blocks:
        raise RuntimeError(f"No step blocks found in xlsx: {value_xlsx_path}")
    return step_blocks


def _parse_steps_arg(steps_arg: Optional[str]) -> Optional[List[int]]:
    if not steps_arg:
        return None
    return [int(x.strip()) for x in steps_arg.split(",") if x.strip()]


def select_steps(
    step_blocks: Sequence[StepBlock],
    steps_arg: Optional[str],
    step_start: Optional[int],
    step_end: Optional[int],
    last_n: Optional[int],
) -> List[StepBlock]:
    selected = list(step_blocks)
    explicit_steps = _parse_steps_arg(steps_arg)
    if explicit_steps is not None:
        step_set = set(explicit_steps)
        selected = [block for block in selected if block.step_idx in step_set]
    if step_start is not None:
        selected = [block for block in selected if block.step_idx >= step_start]
    if step_end is not None:
        selected = [block for block in selected if block.step_idx <= step_end]
    if last_n is not None:
        selected = selected[-int(last_n) :]
    if not selected:
        raise RuntimeError("No steps remain after filtering.")
    return selected


def build_step_weights(step_blocks: Sequence[StepBlock], weight_mode: str) -> np.ndarray:
    num_steps = len(step_blocks)
    if weight_mode == "uniform":
        weights = np.ones(num_steps, dtype=np.float64)
    elif weight_mode == "linear":
        weights = np.arange(1, num_steps + 1, dtype=np.float64)
    elif weight_mode == "confidence_mean":
        weights = np.asarray(
            [max(float(np.mean(block.confidence)), 1e-8) for block in step_blocks],
            dtype=np.float64,
        )
    else:
        raise ValueError(f"Unsupported weight mode: {weight_mode}")
    weights /= weights.sum()
    return weights


def aggregate_attention_graph(
    step_blocks: Sequence[StepBlock],
    weights: np.ndarray,
    confidence_mode: str,
    local_beta: float,
    sym_mode: str,
) -> np.ndarray:
    seq_len = step_blocks[0].attention.shape[0]
    agg = np.zeros((seq_len, seq_len), dtype=np.float64)
    idx = np.arange(seq_len, dtype=np.float64)
    local_gate = np.exp(-float(local_beta) * np.abs(idx[:, None] - idx[None, :]))
    for weight, block in zip(weights, step_blocks):
        mat = np.asarray(block.attention, dtype=np.float64)
        if sym_mode == "avg":
            mat = 0.5 * (mat + mat.T)
        elif sym_mode == "max":
            mat = np.maximum(mat, mat.T)
        else:
            raise ValueError(f"Unsupported sym mode: {sym_mode}")
        conf = np.clip(block.confidence.astype(np.float64), 0.0, 1.0)
        if confidence_mode == "none":
            conf_gate = 1.0
        elif confidence_mode == "outer":
            conf_gate = np.sqrt(np.outer(conf, conf))
        elif confidence_mode == "row":
            conf_gate = conf[:, None]
        else:
            raise ValueError(f"Unsupported confidence mode: {confidence_mode}")
        agg += weight * mat * conf_gate * local_gate
    agg = np.maximum(agg, 0.0)
    np.fill_diagonal(agg, np.diag(agg) + 1e-8)
    return agg


def _clean_token(token: str) -> str:
    return token.split(" (id=")[0]


def _keep_token(token: str) -> bool:
    text = _clean_token(token)
    if any(pattern in text for pattern in SPECIAL_PATTERNS):
        return False
    if text.strip() == "":
        return False
    return True


def build_line_units(tokens: Sequence[str], keep_mask: np.ndarray) -> List[List[int]]:
    units: List[List[int]] = []
    cur: List[int] = []
    for idx, token in enumerate(tokens):
        text = _clean_token(token)
        if text in LINE_BREAK_TOKENS:
            if cur:
                units.append(cur)
                cur = []
            continue
        if not keep_mask[idx]:
            continue
        cur.append(idx)
    if cur:
        units.append(cur)
    if not units:
        raise RuntimeError("No line units remain after filtering.")
    return merge_structural_units(units, tokens)


def is_structural_shell_text(text: str) -> bool:
    text = text.strip()
    if not text:
        return True
    if len(text) <= 8 or text in {"<subtask>", "</subtask>", "{", "}", "}}"}:
        return True
    if "<subtask>" in text and '"' not in text and ":" not in text:
        return True
    if "</subtask>" in text and '"' not in text and ":" not in text:
        return True
    return False


def is_closing_structural_shell_text(text: str) -> bool:
    text = text.strip()
    if not text:
        return False
    if text in {"}", "}}", "</subtask>"}:
        return True
    if "</subtask>" in text and '"' not in text and ":" not in text:
        return True
    if text.endswith("}") and '"' not in text and ":" not in text:
        return True
    return False


def merge_structural_units(units: Sequence[Sequence[int]], tokens: Sequence[str]) -> List[List[int]]:
    merged: List[List[int]] = []
    for unit in units:
        text = "".join(_clean_token(tokens[idx]) for idx in unit).strip()
        is_structural = is_structural_shell_text(text)
        if is_structural and is_closing_structural_shell_text(text) and merged:
            merged[-1].extend(unit)
        else:
            merged.append(list(unit))

    idx = 0
    while idx < len(merged) - 1:
        text = "".join(_clean_token(tokens[token_idx]) for token_idx in merged[idx]).strip()
        if is_structural_shell_text(text) and not is_closing_structural_shell_text(text):
            merged[idx + 1] = merged[idx] + merged[idx + 1]
            del merged[idx]
            continue
        idx += 1
    return merged


def aggregate_units(adj: np.ndarray, seq_indices: np.ndarray, tokens: Sequence[str], units: Sequence[Sequence[int]]):
    n = len(units)
    unit_adj = np.zeros((n, n), dtype=np.float64)
    unit_seq = np.zeros(n, dtype=np.int64)
    unit_texts: List[str] = []
    spans: List[Tuple[int, int]] = []
    for i, group_i in enumerate(units):
        spans.append((group_i[0], group_i[-1]))
        unit_seq[i] = int(seq_indices[group_i[0]])
        unit_texts.append("".join(_clean_token(tokens[idx]) for idx in group_i).strip())
        for j, group_j in enumerate(units):
            block = adj[np.ix_(group_i, group_j)]
            unit_adj[i, j] = float(np.mean(block)) if block.size else 0.0
    return unit_adj, unit_seq, unit_texts, spans


def contiguous_blocks(labels: np.ndarray) -> List[Tuple[int, int, int]]:
    blocks: List[Tuple[int, int, int]] = []
    start = 0
    cur = int(labels[0])
    for idx in range(1, labels.size):
        if int(labels[idx]) != cur:
            blocks.append((start, idx, cur))
            start = idx
            cur = int(labels[idx])
    blocks.append((start, labels.size, cur))
    return blocks


def boundary_attention_score(unit_adj: np.ndarray, boundary: int, window: int) -> float:
    left = np.arange(max(0, boundary - window), boundary)
    right = np.arange(boundary, min(unit_adj.shape[0], boundary + window))
    if left.size == 0 or right.size == 0:
        return 0.0
    return float(np.mean(unit_adj[np.ix_(left, right)]))


def boundary_structure_bonus(prev_text: str, next_text: str) -> float:
    prev_text = prev_text.strip()
    next_text = next_text.strip()
    bonus = 0.0
    if "</subtask>" in prev_text or "}}" in prev_text or prev_text.endswith("},") or prev_text == "}":
        bonus += 1.0
    if "<subtask>" in next_text or next_text.startswith("<sub") or "subtask_name" in next_text:
        bonus += 1.0
    return bonus


def hybrid_boundary_scores(unit_adj: np.ndarray, unit_texts: Sequence[str], window: int, structure_weight: float) -> np.ndarray:
    scores = np.zeros(unit_adj.shape[0] - 1, dtype=np.float64)
    for boundary in range(1, unit_adj.shape[0]):
        attn = boundary_attention_score(unit_adj, boundary, window)
        structure = boundary_structure_bonus(unit_texts[boundary - 1], unit_texts[boundary])
        scores[boundary - 1] = -attn + structure_weight * structure
    return scores


def choose_boundaries(scores: np.ndarray, k: int, min_block_len: int) -> List[int]:
    if k <= 1:
        return []
    ranked = sorted(range(scores.shape[0]), key=lambda idx: float(scores[idx]), reverse=True)
    selected: List[int] = []
    for idx in ranked:
        boundary = idx + 1
        if any(abs(boundary - prev) < min_block_len for prev in selected):
            continue
        selected.append(boundary)
        if len(selected) >= max(0, k - 1):
            break
    return sorted(selected)


def evaluate_k(
    unit_adj: np.ndarray,
    unit_texts: Sequence[str],
    min_k: int,
    max_k: int,
    window: int,
    structure_weight: float,
    split_penalty: float,
    min_block_len: int,
) -> List[Tuple[int, float, np.ndarray, List[int]]]:
    boundary_scores = hybrid_boundary_scores(unit_adj, unit_texts, window, structure_weight)
    results: List[Tuple[int, float, np.ndarray, List[int]]] = []
    for k in range(max(1, min_k), max(min_k, max_k) + 1):
        boundaries = choose_boundaries(boundary_scores, k, min_block_len)
        labels = np.zeros(unit_adj.shape[0], dtype=np.int64)
        points = [0] + boundaries + [unit_adj.shape[0]]
        for label, (start, end) in enumerate(zip(points[:-1], points[1:])):
            labels[start:end] = label
        score = sum(boundary_scores[b - 1] for b in boundaries) - split_penalty * max(0, len(boundaries))
        results.append((k, score, labels, boundaries))
    return results


def write_matrix(ws, matrix: np.ndarray, seq_indices: np.ndarray) -> None:
    ws.append(["unit_pos", "seq_idx"] + [f"u_{i + 1}" for i in range(matrix.shape[1])])
    for i in range(matrix.shape[0]):
        ws.append([i + 1, int(seq_indices[i])] + matrix[i, :].astype(float).tolist())


def render_block_text(unit_texts: Sequence[str], start: int, end: int) -> str:
    return "\n".join(unit_texts[start:end]).strip()


def save_block_text(output_path: str, unit_texts: Sequence[str], blocks: Sequence[Tuple[int, int, int]]) -> None:
    lines: List[str] = []
    for block_id, (start, end, label) in enumerate(blocks, start=1):
        lines.append(f"[Block {block_id}] label={label} units={start + 1}-{end}")
        lines.append(render_block_text(unit_texts, start, end))
        lines.append("")
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines).rstrip() + "\n")


def block_color(label: int) -> str:
    palette = [
        "#dbeafe",
        "#dcfce7",
        "#fef3c7",
        "#fce7f3",
        "#ede9fe",
        "#ffe4e6",
        "#e0f2fe",
        "#ecfccb",
    ]
    return palette[int(label) % len(palette)]


def _mix_hex_colors(base_hex: str, accent_hex: str, weight: float) -> str:
    weight = max(0.0, min(1.0, float(weight)))
    base = tuple(int(base_hex[i : i + 2], 16) for i in (1, 3, 5))
    accent = tuple(int(accent_hex[i : i + 2], 16) for i in (1, 3, 5))
    mixed = tuple(int(round((1.0 - weight) * b + weight * a)) for b, a in zip(base, accent))
    return "#" + "".join(f"{channel:02x}" for channel in mixed)


def build_heatmap_html(
    unit_adj: np.ndarray,
    labels: np.ndarray,
    boundaries: Sequence[int],
    blocks: Sequence[Tuple[int, int, int]],
) -> str:
    vmax = float(np.max(unit_adj))
    scale = vmax if vmax > 1e-12 else 1.0
    boundary_after = set(int(x) for x in boundaries)
    block_ranges = {int(label): (int(start), int(end)) for start, end, label in blocks}
    rows: List[str] = []
    for i in range(unit_adj.shape[0]):
        row_cells: List[str] = []
        for j in range(unit_adj.shape[1]):
            value = float(unit_adj[i, j])
            norm = max(0.0, min(1.0, value / scale))
            same_block = labels[i] == labels[j]
            base_color = "#f8fafc" if not same_block else block_color(int(labels[i]))
            fill_color = _mix_hex_colors(base_color, "#0f766e", 0.12 + 0.78 * norm)

            classes = ["heat"]
            if not same_block:
                classes.append("cross-block")
            if (j + 1) in boundary_after:
                classes.append("boundary-col")
            if (i + 1) in boundary_after:
                classes.append("boundary-row")

            start, end = block_ranges[int(labels[i])]
            styles = [f"background: {fill_color}"]
            if same_block and i == start:
                styles.append("border-top: 3px solid #111827")
            if same_block and i == end - 1:
                styles.append("border-bottom: 3px solid #111827")
            if same_block and j == start:
                styles.append("border-left: 3px solid #111827")
            if same_block and j == end - 1:
                styles.append("border-right: 3px solid #111827")

            row_cells.append(
                f'<td class="{" ".join(classes)}" style="{"; ".join(styles)}" '
                f'title="u{i + 1} ↔ u{j + 1} | value={value:.6g} | same_block={same_block}">{value:.3g}</td>'
            )
        rows.append("<tr>" + "".join(row_cells) + "</tr>")
    return "\n".join(rows)


def save_html_report(
    output_path: str,
    unit_adj: np.ndarray,
    unit_seq: np.ndarray,
    unit_texts: Sequence[str],
    spans: Sequence[Tuple[int, int]],
    labels: np.ndarray,
    boundaries: Sequence[int],
    blocks: Sequence[Tuple[int, int, int]],
) -> None:
    unit_rows: List[str] = []
    boundary_set = set(int(x) for x in boundaries)
    for idx, (seq_idx, text, span, label) in enumerate(zip(unit_seq, unit_texts, spans, labels), start=1):
        marker = "Yes" if idx in boundary_set else ""
        label_bg = block_color(int(label))
        unit_rows.append(
            "<tr>"
            f"<td>{idx}</td>"
            f'<td style="background: {label_bg}; font-weight: 700;">{int(label)}</td>'
            f"<td>{marker}</td>"
            f"<td>{int(seq_idx)}</td>"
            f"<td>{span[0] + 1}-{span[1] + 1}</td>"
            f"<td><pre>{html.escape(text)}</pre></td>"
            "</tr>"
        )

    block_cards: List[str] = []
    for block_id, (start, end, label) in enumerate(blocks, start=1):
        text = render_block_text(unit_texts, start, end)
        block_cards.append(
            f'<section class="block-card" style="border-top: 8px solid {block_color(int(label))};">'
            f"<h3>Block {block_id}</h3>"
            f"<p>label={int(label)} | units={start + 1}-{end} | tokens={spans[start][0] + 1}-{spans[end - 1][1] + 1}</p>"
            f"<pre>{html.escape(text)}</pre>"
            "</section>"
        )

    header_cells = "".join(f"<th>{i}</th>" for i in range(1, unit_adj.shape[0] + 1))
    heatmap_rows = build_heatmap_html(unit_adj, labels, boundaries, blocks)
    row_headers = "\n".join(
        f'<tr><th style="background: {block_color(int(labels[i]))};">{i + 1}</th>{row}</tr>'
        for i, row in enumerate(heatmap_rows.splitlines())
    )
    legend_items = "".join(
        f'<span class="legend-item"><span class="legend-swatch" style="background: {block_color(int(label))};"></span>'
        f'Block {block_id} (label={int(label)}, units={start + 1}-{end})</span>'
        for block_id, (start, end, label) in enumerate(blocks, start=1)
    )
    boundary_text = ", ".join(str(x) for x in boundaries) or "None"

    html_doc = f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <title>Response Block Visualization</title>
  <style>
    body {{ font-family: ui-monospace, SFMono-Regular, Menlo, monospace; margin: 24px; color: #17212b; background: #f6f8fb; }}
    h1, h2, h3 {{ margin: 0 0 12px; }}
    section {{ margin: 0 0 28px; }}
    table {{ border-collapse: collapse; background: white; }}
    th, td {{ border: 1px solid #d8e0ea; padding: 6px 8px; vertical-align: top; }}
    th {{ background: #eef3f8; position: sticky; top: 0; }}
    pre {{ white-space: pre-wrap; word-break: break-word; margin: 0; }}
    .matrix-wrap {{ overflow: auto; max-width: 100%; border: 1px solid #d8e0ea; background: white; box-shadow: 0 8px 24px rgba(15, 23, 42, 0.06); }}
    .heat {{ min-width: 48px; text-align: right; font-size: 12px; transition: transform 0.08s ease; }}
    .heat:hover {{ transform: scale(1.08); position: relative; z-index: 2; }}
    .cross-block {{ color: #475569; }}
    .boundary-row {{ border-top: 3px dashed #dc2626 !important; }}
    .boundary-col {{ border-left: 3px dashed #dc2626 !important; }}
    .grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(320px, 1fr)); gap: 16px; }}
    .block-card {{ background: white; border: 1px solid #d8e0ea; padding: 16px; box-shadow: 0 8px 24px rgba(15, 23, 42, 0.05); }}
    .meta {{ background: white; border: 1px solid #d8e0ea; padding: 12px 16px; }}
    .legend {{ display: flex; flex-wrap: wrap; gap: 10px 16px; margin-top: 12px; }}
    .legend-item {{ display: inline-flex; align-items: center; gap: 8px; background: #f8fafc; border: 1px solid #d8e0ea; padding: 6px 10px; }}
    .legend-swatch {{ width: 18px; height: 18px; display: inline-block; border: 1px solid #94a3b8; }}
    .note {{ color: #475569; margin-top: 10px; }}
  </style>
</head>
<body>
  <section class="meta">
    <h1>Response Block Visualization</h1>
    <p>Units: {len(unit_texts)} | Blocks: {len(blocks)} | Boundaries after units: {boundary_text}</p>
    <div class="legend">{legend_items}</div>
    <p class="note">Heat color gets stronger as unit-to-unit affinity increases. Dashed red guides mark chosen cut boundaries. Thick dark square edges outline each final block on the heatmap diagonal.</p>
  </section>
  <section>
    <h2>Unit Graph Heatmap</h2>
    <div class="matrix-wrap">
      <table>
        <tr><th>u</th>{header_cells}</tr>
        {row_headers}
      </table>
    </div>
  </section>
  <section>
    <h2>Units</h2>
    <table>
      <tr><th>Unit</th><th>Label</th><th>Boundary After</th><th>Seq Idx</th><th>Token Span</th><th>Text</th></tr>
      {"".join(unit_rows)}
    </table>
  </section>
  <section>
    <h2>Blocks</h2>
    <div class="grid">
      {"".join(block_cards)}
    </div>
  </section>
</body>
</html>
"""
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html_doc)


def save_results(
    output_path: str,
    unit_adj: np.ndarray,
    unit_seq: np.ndarray,
    unit_texts: Sequence[str],
    spans: Sequence[Tuple[int, int]],
    k_results: Sequence[Tuple[int, float, np.ndarray, List[int]]],
    chosen_k: int,
) -> None:
    chosen = next(item for item in k_results if item[0] == chosen_k)
    labels = chosen[2]
    boundaries = chosen[3]
    blocks = contiguous_blocks(labels)

    wb = Workbook()
    ws_adj = wb.active
    ws_adj.title = "unit_graph"
    write_matrix(ws_adj, unit_adj, unit_seq)

    ws_units = wb.create_sheet("units")
    ws_units.append(["unit_pos", "seq_idx", "start_token_pos", "end_token_pos", "label", "text"])
    for idx, (seq_idx, text, span, label) in enumerate(zip(unit_seq, unit_texts, spans, labels), start=1):
        ws_units.append([idx, int(seq_idx), span[0] + 1, span[1] + 1, int(label), text])

    ws_k = wb.create_sheet("k_candidates")
    ws_k.append(["k", "score_higher_is_better"])
    for k, score, _, _ in k_results:
        ws_k.append([int(k), float(score)])

    ws_block = wb.create_sheet("response_blocks")
    ws_block.append(["block_id", "start_unit", "end_unit", "label", "start_token_pos", "end_token_pos"])
    for block_id, (start, end, label) in enumerate(blocks, start=1):
        ws_block.append([block_id, start + 1, end, int(label), spans[start][0] + 1, spans[end - 1][1] + 1])

    ws_meta = wb.create_sheet("meta")
    ws_meta.append(["chosen_k", int(chosen_k)])
    ws_meta.append(["chosen_boundaries", ",".join(str(x) for x in boundaries)])
    ws_meta.append(["num_units", int(len(unit_texts))])
    ws_meta.append(["num_blocks_after_smoothing", int(len(blocks))])

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)
    base, _ = os.path.splitext(output_path)
    save_html_report(
        output_path=base + ".html",
        unit_adj=unit_adj,
        unit_seq=unit_seq,
        unit_texts=unit_texts,
        spans=spans,
        labels=labels,
        boundaries=boundaries,
        blocks=blocks,
    )
    save_block_text(base + "_blocks.txt", unit_texts, blocks)


def main() -> None:
    parser = argparse.ArgumentParser(description="Detect response blocks with structure-aware unit clustering.")
    parser.add_argument("--value-xlsx", type=str, default="/data/home/yzx/LLaDA/yzx_test/value_outputs/denoise_value.xlsx")
    parser.add_argument("--value-sheet", type=str, default="denoise_values_mask_mask")
    parser.add_argument("--output-xlsx", type=str, default="/data/home/yzx/LLaDA/yzx_test/value_outputs/blocks_structure.xlsx")
    parser.add_argument("--steps", type=str, default=None)
    parser.add_argument("--step-start", type=int, default=None)
    parser.add_argument("--step-end", type=int, default=None)
    parser.add_argument("--last-n", type=int, default=8)
    parser.add_argument("--weight-mode", type=str, default="confidence_mean", choices=["uniform", "linear", "confidence_mean"])
    parser.add_argument("--confidence-mode", type=str, default="outer", choices=["none", "outer", "row"])
    parser.add_argument("--sym-mode", type=str, default="avg", choices=["avg", "max"])
    parser.add_argument("--local-beta", type=float, default=0.003)
    parser.add_argument("--min-k", type=int, default=1)
    parser.add_argument("--max-k", type=int, default=6)
    parser.add_argument("--window", type=int, default=2)
    parser.add_argument("--structure-weight", type=float, default=5.0)
    parser.add_argument("--split-penalty", type=float, default=1.0)
    parser.add_argument("--min-block-len", type=int, default=2)
    parser.add_argument("--force-k", type=int, default=None)
    args = parser.parse_args()

    step_blocks = load_step_blocks(args.value_xlsx, args.value_sheet)
    step_blocks = select_steps(step_blocks, args.steps, args.step_start, args.step_end, args.last_n)
    weights = build_step_weights(step_blocks, args.weight_mode)
    agg_adj = aggregate_attention_graph(step_blocks, weights, args.confidence_mode, args.local_beta, args.sym_mode)

    tokens = step_blocks[0].tokens
    seq_indices = step_blocks[0].seq_indices
    keep_mask = np.asarray([_keep_token(token) for token in tokens], dtype=bool)
    units = build_line_units(tokens, keep_mask)
    unit_adj, unit_seq, unit_texts, spans = aggregate_units(agg_adj, seq_indices, tokens, units)

    k_results = evaluate_k(
        unit_adj,
        unit_texts,
        args.min_k,
        args.max_k,
        args.window,
        args.structure_weight,
        args.split_penalty,
        args.min_block_len,
    )
    chosen_k = int(args.force_k) if args.force_k is not None else max(k_results, key=lambda item: item[1])[0]
    save_results(args.output_xlsx, unit_adj, unit_seq, unit_texts, spans, k_results, chosen_k)
    chosen = next(item for item in k_results if item[0] == chosen_k)
    print(f"Saved structure-aware workbook to: {args.output_xlsx}")
    print(f"Saved HTML report to: {os.path.splitext(args.output_xlsx)[0] + '.html'}")
    print(f"Saved block text to: {os.path.splitext(args.output_xlsx)[0] + '_blocks.txt'}")
    print(f"Chosen K: {chosen_k}")
    print(f"Blocks after smoothing: {len(contiguous_blocks(chosen[2]))}")


if __name__ == "__main__":
    main()
